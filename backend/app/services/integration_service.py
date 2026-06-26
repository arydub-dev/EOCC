"""Data Integration Center service: connectors, CSV/Excel import, pipeline."""

from __future__ import annotations

import csv
import io
import time
from datetime import UTC, datetime

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import settings
from app.models import DataSource, Hospital, ImportJob, Incident, Resource, Shelter, enums
from app.schemas.integration import IntegrationOverview, PipelineMonitor
from app.services import file_security, scoring_service
from app.services.file_security import FileSecurityError, sanitize_cell


def overview(db: Session) -> IntegrationOverview:
    sources = list(db.scalars(select(DataSource)).all())
    healthy = sum(1 for s in sources if s.status == enums.DataSourceStatus.HEALTHY)
    degraded = sum(1 for s in sources if s.status == enums.DataSourceStatus.DEGRADED)
    offline = sum(
        1
        for s in sources
        if s.status in (enums.DataSourceStatus.OFFLINE, enums.DataSourceStatus.DISCONNECTED)
    )
    total_records = sum(s.records_synced for s in sources)
    avg_health = round(sum(s.health_score for s in sources) / len(sources), 1) if sources else 0.0
    last_sync = max((s.last_sync_at for s in sources if s.last_sync_at), default=None)
    return IntegrationOverview(
        connected_systems=len(sources),
        healthy=healthy,
        degraded=degraded,
        offline=offline,
        total_records_synced=total_records,
        avg_health_score=avg_health,
        last_sync_at=last_sync,
        sources=sources,  # type: ignore[arg-type]
    )


def pipeline_monitor(db: Session) -> PipelineMonitor:
    jobs = list(db.scalars(select(ImportJob).order_by(ImportJob.created_at.desc())).all())
    completed = sum(1 for j in jobs if j.status == enums.ImportJobStatus.COMPLETED)
    failed = sum(1 for j in jobs if j.status == enums.ImportJobStatus.FAILED)
    partial = sum(1 for j in jobs if j.status == enums.ImportJobStatus.PARTIAL)
    processed = sum(j.records_processed for j in jobs)
    rec_failed = sum(j.records_failed for j in jobs)
    durations = [j.duration_ms for j in jobs if j.duration_ms]
    avg_duration = round(sum(durations) / len(durations), 1) if durations else 0.0
    total = len(jobs)
    health = round((completed / total * 100) if total else 100.0, 1)
    return PipelineMonitor(
        total_jobs=total,
        completed=completed,
        failed=failed,
        partial=partial,
        records_processed=processed,
        records_failed=rec_failed,
        avg_duration_ms=avg_duration,
        pipeline_health=health,
        recent_jobs=jobs[:15],  # type: ignore[arg-type]
    )


# ── Import handlers ──
def _parse_float(value: str, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _parse_int(value: str, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _import_hospital(db: Session, row: dict, org_id: int | None, origin: enums.DataOrigin) -> None:
    h = Hospital(
        organization_id=org_id,
        data_origin=origin,
        name=row.get("name", "Imported Hospital"),
        region=row.get("region"),
        latitude=_parse_float(row.get("latitude", "0")),
        longitude=_parse_float(row.get("longitude", "0")),
        total_beds=_parse_int(row.get("total_beds", "0")),
        occupied_beds=_parse_int(row.get("occupied_beds", "0")),
        icu_beds=_parse_int(row.get("icu_beds", "0")),
        icu_occupied=_parse_int(row.get("icu_occupied", "0")),
        er_capacity=_parse_int(row.get("er_capacity", "0")),
        er_patients=_parse_int(row.get("er_patients", "0")),
        staff_on_duty=_parse_int(row.get("staff_on_duty", "0")),
        staff_required=_parse_int(row.get("staff_required", "0")),
        ventilators_total=_parse_int(row.get("ventilators_total", "0")),
        ventilators_in_use=_parse_int(row.get("ventilators_in_use", "0")),
    )
    scoring_service.recompute_hospital(h)
    db.add(h)


def _import_shelter(db: Session, row: dict, org_id: int | None, origin: enums.DataOrigin) -> None:
    sh = Shelter(
        organization_id=org_id,
        data_origin=origin,
        name=row.get("name", "Imported Shelter"),
        region=row.get("region"),
        latitude=_parse_float(row.get("latitude", "0")),
        longitude=_parse_float(row.get("longitude", "0")),
        capacity=_parse_int(row.get("capacity", "0")),
        occupancy=_parse_int(row.get("occupancy", "0")),
        food_supply_days=_parse_float(row.get("food_supply_days", "0")),
        water_supply_days=_parse_float(row.get("water_supply_days", "0")),
        medical_kits=_parse_int(row.get("medical_kits", "0")),
        medical_staff=_parse_int(row.get("medical_staff", "0")),
    )
    scoring_service.recompute_shelter(sh)
    db.add(sh)


def _import_resource(db: Session, row: dict, org_id: int | None, origin: enums.DataOrigin) -> None:
    r = Resource(
        organization_id=org_id,
        data_origin=origin,
        name=row.get("name", "Imported Resource"),
        resource_type=enums.ResourceType(row.get("resource_type", "ambulance")),
        region=row.get("region"),
        latitude=_parse_float(row.get("latitude", "0")),
        longitude=_parse_float(row.get("longitude", "0")),
        capacity=_parse_float(row.get("capacity", "0")),
        quantity_available=_parse_float(row.get("quantity_available", "0")),
        readiness=_parse_float(row.get("readiness", "100")),
    )
    db.add(r)


def _import_incident(db: Session, row: dict, org_id: int | None, origin: enums.DataOrigin) -> None:
    inc = Incident(
        organization_id=org_id,
        data_origin=origin,
        name=row.get("name", "Imported Incident"),
        incident_type=enums.IncidentType(row.get("incident_type", "flood")),
        region=row.get("region"),
        latitude=_parse_float(row.get("latitude", "0")),
        longitude=_parse_float(row.get("longitude", "0")),
        radius_km=_parse_float(row.get("radius_km", "5")),
        severity=_parse_int(row.get("severity", "3")),
        population_impacted=_parse_int(row.get("population_impacted", "0")),
        started_at=datetime.now(UTC),
    )
    scoring_service.recompute_incident(inc)
    db.add(inc)


_IMPORTERS = {
    "hospitals": _import_hospital,
    "shelters": _import_shelter,
    "resources": _import_resource,
    "incidents": _import_incident,
}


def run_csv_import(
    db: Session,
    target_entity: str,
    content: str,
    filename: str | None,
    source_type: enums.DataSourceType = enums.DataSourceType.CSV_IMPORT,
    org_id: int | None = None,
) -> ImportJob:
    importer = _IMPORTERS.get(target_entity)
    origin = (
        enums.DataOrigin.EXCEL
        if source_type == enums.DataSourceType.EXCEL_IMPORT
        else enums.DataOrigin.CSV
    )
    job = ImportJob(
        organization_id=org_id,
        source_type=source_type,
        target_entity=target_entity,
        status=enums.ImportJobStatus.RUNNING,
        filename=filename,
        started_at=datetime.now(UTC),
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    started = time.perf_counter()
    errors: list[dict] = []
    processed = failed = total = 0

    if importer is None:
        job.status = enums.ImportJobStatus.FAILED
        job.errors = [{"row": 0, "error": f"Unsupported target entity: {target_entity}"}]
        job.finished_at = datetime.now(UTC)
        db.commit()
        db.refresh(job)
        return job

    reader = csv.DictReader(io.StringIO(content))
    # Schema validation: reject the whole file early if required columns missing.
    try:
        file_security.validate_schema(target_entity, list(reader.fieldnames or []))
    except FileSecurityError as exc:
        job.status = enums.ImportJobStatus.FAILED
        job.errors = [{"row": 0, "error": str(exc)}]
        job.finished_at = datetime.now(UTC)
        db.commit()
        db.refresh(job)
        return job

    for idx, row in enumerate(reader, start=1):
        total += 1
        if total > settings.MAX_IMPORT_ROWS:
            errors.append({"row": idx, "error": "Row limit exceeded; import truncated."})
            break
        try:
            # Strip whitespace and neutralize CSV/formula injection on every cell.
            clean = {k.strip(): sanitize_cell((v or "").strip()) for k, v in row.items() if k}
            importer(db, clean, org_id, origin)
            processed += 1
        except Exception as exc:  # noqa: BLE001 - report row-level failures
            failed += 1
            errors.append({"row": idx, "error": str(exc)})

    db.commit()
    job.records_total = total
    job.records_processed = processed
    job.records_failed = failed
    job.errors = errors or None
    job.duration_ms = int((time.perf_counter() - started) * 1000)
    job.finished_at = datetime.now(UTC)
    if failed == 0:
        job.status = enums.ImportJobStatus.COMPLETED
    elif processed == 0:
        job.status = enums.ImportJobStatus.FAILED
    else:
        job.status = enums.ImportJobStatus.PARTIAL
    db.commit()
    db.refresh(job)
    return job


def run_excel_import(
    db: Session,
    target_entity: str,
    file_bytes: bytes,
    filename: str | None,
    org_id: int | None = None,
) -> ImportJob:
    """Convert the first worksheet to CSV text, then reuse the CSV importer."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(["" if cell is None else cell for cell in row])
    return run_csv_import(
        db,
        target_entity,
        buffer.getvalue(),
        filename,
        source_type=enums.DataSourceType.EXCEL_IMPORT,
        org_id=org_id,
    )
